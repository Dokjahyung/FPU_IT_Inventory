from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

wb1 = load_workbook('/')
wb2 = load_workbook('/')


# Access the active sheet
sheet = wb1.active
sheet2 = wb2.active
# Modify the content of a specific cell (e.g., A1)
def apply_formatting(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
            cell.font = Font(color="000000")

def dictc(sheet):
    data_dict = {}
    for row in sheet.iter_rows(min_row=0, values_only=True):
        key = row[0]  # Assuming the first column is the key
        value = [row[1],row[2],row[3]]  # Assuming the second column is the value
        data_dict[key] = value
    return data_dict

apply_formatting(sheet)
apply_formatting(sheet2)
sheet2_data = dictc(sheet2)
print(sheet2_data)

sheet1_data = {}
for row in sheet.iter_rows(min_row=0, values_only=True):
    key = row[0]  # Assuming the first column is the key
    value = [row[10], f"{row[1]} {row[2]}", f"{row[3]} {row[5]}"]  # Assuming the second column is the value
    sheet1_data[key] = value
# Save the modified workbook





def sort(sheet, sheet1_data, sheet2_data):
    current_row = 1  # Assuming headers are in the first row
    for row in sheet.iter_rows(min_row=current_row):
        key = row[0].value
        if key in sheet1_data and key in sheet2_data:
            row[0].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            if sheet1_data[key] == sheet2_data[key]:
                for cell in row:
                    cell.fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            if sheet1_data[key][0] == sheet2_data[key][0]:
                row[10].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            else:
                print(str(row[0].value), "wrong device")
                row[10].fill = PatternFill(start_color="FEF5A8", end_color="FEF5A8", fill_type="solid")
            if sheet1_data[key][1] is not None and sheet2_data[key][1] is not None and sheet2_data[key][1] in sheet1_data[key][1]:
                row[1].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
                row[2].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            else:
                print(str(row[0].value), "wrong user")
                row[1].fill = PatternFill(start_color="FEF5A8", end_color="FEF5A8", fill_type="solid")
                row[2].fill = PatternFill(start_color="FEF5A8", end_color="FEF5A8", fill_type="solid")
            if sheet1_data[key][2] == sheet2_data[key][2]:
                row[3].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
                row[4].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
                row[5].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            else:
                print(str(row[0].value), "wrong room")
                row[3].fill = PatternFill(start_color="FEF5A8", end_color="FEF5A8", fill_type="solid")
                row[4].fill = PatternFill(start_color="FEF5A8", end_color="FEF5A8", fill_type="solid")
                row[5].fill = PatternFill(start_color="FEF5A8", end_color="FEF5A8", fill_type="solid")
        else:
            for cell in row:
                    cell.fill = PatternFill(start_color="F79494", end_color="F79494", fill_type="solid")
        current_row += 1




def sort2(sheet, sheet1_data, sheet2_data):
    current_row = 1  # Assuming headers are in the first row
    for row in sheet.iter_rows(min_row=current_row):
        key = row[0].value
        if key in sheet1_data and key in sheet2_data:
            row[0].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            if sheet1_data[key] == sheet2_data[key]:
                for cell in row:
                    cell.fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            if sheet1_data[key][0] == sheet2_data[key][0]:
                row[1].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            else:
                print(str(row[0].value), "wrong device")
                row[1].fill = PatternFill(start_color="FEF5A8", end_color="FEF5A8", fill_type="solid")
            if sheet1_data[key][1] is not None and sheet2_data[key][1] is not None and sheet2_data[key][1] in sheet1_data[key][1]:
                row[2].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            else:
                print(str(row[0].value), "wrong user")
                row[2].fill = PatternFill(start_color="FEF5A8", end_color="FEF5A8", fill_type="solid")
            if sheet1_data[key][2] == sheet2_data[key][2]:
                row[3].fill = PatternFill(start_color="c1f9a2", end_color="c1f9a2", fill_type="solid")
            else:
                print(str(row[0].value), "wrong room")
                row[3].fill = PatternFill(start_color="FEF5A8", end_color="FEF5A8", fill_type="solid")
        else:
            for cell in row:
                    cell.fill = PatternFill(start_color="F79494", end_color="F79494", fill_type="solid")
        current_row += 1





sort2(sheet2, sheet1_data, sheet2_data)
sort(sheet, sheet1_data, sheet2_data)
print(str(171), sheet1_data[171])
wb1.save('/')
wb2.save('/')
print('finish')