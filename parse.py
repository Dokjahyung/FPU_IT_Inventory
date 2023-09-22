from openpyxl import load_workbook

wb1 = load_workbook('/')
sheet = wb1.active

def parse(sheet):
    for row in sheet.iter_rows(min_row=2):  # Assuming headers are in the first row
        device_description = row[10].value
        cell_select = row[10]

        match device_description:
            case "Monitor - Flat Panel":
                cell_select.value = "M"
                
            case "Laptop Computer":
                cell_select.value = "L"
                
            case "Desktop Computer":
                cell_select.value = "PC"
                
            case _:
                cell_select.value = device_description  # Set to the original description for other cases

parse(sheet)

wb1.save('/')
